"""Document text extraction for PDF, DOCX, TXT, MD, CSV, and Excel."""
from typing import Optional
import io


class DocumentParser:
    @staticmethod
    def extract_text(content_bytes: bytes, filename: str) -> str:
        lower = filename.lower()
        if lower.endswith('.pdf'):
            try:
                from PyPDF2 import PdfReader
                reader = PdfReader(io.BytesIO(content_bytes))
                pages = []
                for page in reader.pages:
                    text = page.extract_text()
                    if text:
                        pages.append(text)
                return "\n\n".join(pages) or "No text could be extracted from PDF."
            except Exception as e:
                return f"PDF extraction error: {str(e)}"

        elif lower.endswith('.docx'):
            try:
                from docx import Document
                doc = Document(io.BytesIO(content_bytes))
                paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
                return "\n".join(paragraphs) or "No text could be extracted from DOCX."
            except Exception as e:
                return f"DOCX extraction error: {str(e)}"

        elif lower.endswith('.txt') or lower.endswith('.md'):
            return content_bytes.decode('utf-8', errors='ignore')

        elif lower.endswith('.csv'):
            try:
                return content_bytes.decode('utf-8', errors='ignore')
            except Exception as e:
                return f"CSV read error: {str(e)}"

        elif lower.endswith('.xlsx') or lower.endswith('.xls'):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content_bytes), data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    rows = []
                    for row in sheet.iter_rows(values_only=True):
                        row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                        if row_text.strip():
                            rows.append(row_text)
                    if rows:
                        sheets.append(f"--- Sheet: {sheet_name} ---\n" + "\n".join(rows))
                return "\n\n".join(sheets) or "No text could be extracted from Excel file."
            except ImportError:
                return (
                    "Excel parsing requires 'openpyxl'. Install it with:\n"
                    "pip install openpyxl\n\n"
                    "Alternatively, convert your file to CSV and upload that."
                )
            except Exception as e:
                return f"Excel extraction error: {str(e)}"

        else:
            return (
                f"Unsupported file format: {filename}.\n"
                f"Supported formats: PDF, DOCX, TXT, MD, CSV, XLSX, XLS."
            )